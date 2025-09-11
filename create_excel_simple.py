#!/usr/bin/env python3
"""
Create Excel file using openpyxl directly (no pandas required)
"""
import openpyxl
from openpyxl import Workbook

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Shipments"

# Add headers
headers = ['title', 'destination', 'eta', 'status']
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# Add sample data
data = [
    ['Electronics Package', 'New York NY', '2024-01-15', 'pending'],
    ['Office Supplies', 'Los Angeles CA', '2024-01-20', 'agreed'],
    ['Medical Equipment', 'Chicago IL', '2024-01-25', 'pending'],
    ['Furniture Delivery', 'Miami FL', '2024-01-30', 'pending'],
    ['Books and Media', 'Seattle WA', '2024-02-05', 'agreed'],
    ['Automotive Parts', 'Detroit MI', '2024-02-10', 'pending'],
    ['Clothing Shipment', 'Atlanta GA', '2024-02-15', 'agreed'],
    ['Food Products', 'Denver CO', '2024-02-20', 'pending']
]

# Add data rows
for row, row_data in enumerate(data, 2):
    for col, value in enumerate(row_data, 1):
        ws.cell(row=row, column=col, value=value)

# Save the file
wb.save('backend/data/seed_shipments.xlsx')
print("✅ Created backend/data/seed_shipments.xlsx with sample data")
print(f"📊 Created {len(data)} sample shipments")
